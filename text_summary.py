import pandas as pd
import openpyxl
import os
import pyexcel as p
from openpyxl.utils import get_column_letter

# # --- セル値を抽出するヘルパー関数 ---
# def _get_cell_value(worksheet, cell_address):
#     """
#     指定されたワークシートから特定のセル（結合セル対応）の値を抽出します。
#     """
#     try:
#         target_cell = worksheet[cell_address]
        
#         is_merged = False
#         for merged_range in worksheet.merged_cells.ranges:
#             if cell_address in merged_range: 
#                 is_merged = True
#                 top_left_cell_address = merged_range.coord.split(':')[0]
#                 # print(f"  セル '{cell_address}' は結合セル '{merged_range.coord}' の一部です。")
#                 # print(f"  結合元のセル '{top_left_cell_address}' の値を参照します。")
#                 return worksheet[top_left_cell_address].value
        
#         if not is_merged:
#             # print(f"  セル '{cell_address}' は結合セルではないか、結合元のセルです。")
#             return target_cell.value
            
#     except Exception as e:
#         print(f"警告: セル '{cell_address}' の処理中にエラーが発生しました: {e}")
#         return None # エラーの場合はNoneを返す

# # --- 全てのシートから辞書形式で指定された複数のセルペアを抽出する関数 (変更点あり) ---
# def extract_specified_cell_pairs_from_all_sheets(excel_file_path, cell_address_map):
#     """
#     Excelファイルの全てのシートから、辞書形式で指定された複数のセルペアの値を抽出します。

#     Args:
#         excel_file_path (str): 処理するExcelファイルの完全なパス（.xlsx形式である必要があります）。
#         cell_address_map (dict): 抽出したいセルのペアを定義する辞書。
#                                  キーと値がそれぞれ抽出したいセルアドレスとなる。
#                                  例: {'A8': 'V8', 'A10': 'V10'}

#     Returns:
#         dict: 各シート名をキーとし、そのシートから抽出されたセルの値を格納した辞書。
#               セルの値は、元のセルアドレスをキーとする別の辞書として格納されます。
#               例: {
#                   'Sheet1': {'A8': 'A8の値', 'V8': 'V8の値', 'A10': 'A10の値', 'V10': 'V10の値'},
#                   'Sheet2': {'A8': 'A8の値2', 'V8': 'V8の値2', 'A10': 'A10の値2', 'V10': 'V10の値2'}
#               }
#               ファイルが見つからない場合やエラーの場合は空の辞書を返します。
#     """
#     extracted_data_by_sheet = {}

#     if not os.path.exists(excel_file_path):
#         print(f"エラー: Excelファイル '{excel_file_path}' が見つかりません。")
#         return {}

#     if not excel_file_path.lower().endswith('.xlsx'):
#         print(f"エラー: 抽出対象のファイル '{excel_file_path}' は.xlsx形式ではありません。")
#         print("この関数は.xlsxファイルのみをサポートします。")
#         return {}

#     try:
#         workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
        
#         for sheet_name in workbook.sheetnames:
#             print(f"\n--- シート '{sheet_name}' を処理中 ---")
#             ws = workbook[sheet_name]
            
#             # このシートの抽出結果を格納する辞書
#             sheet_extracted_cells = {} 

#             # cell_address_map のキーと値を両方抽出
#             for primary_cell, secondary_cell in cell_address_map.items():
                
#                 # primary_cell (例: A8) の値を取得
#                 value_primary = _get_cell_value(ws, primary_cell)
#                 sheet_extracted_cells[primary_cell] = value_primary
#                 print(f"  セル '{primary_cell}' の値: {value_primary}")

#                 # secondary_cell (例: V8) の値を取得
#                 value_secondary = _get_cell_value(ws, secondary_cell)
#                 sheet_extracted_cells[secondary_cell] = value_secondary
#                 print(f"  セル '{secondary_cell}' の値: {value_secondary}")
            
#             extracted_data_by_sheet[sheet_name] = sheet_extracted_cells

#     except Exception as e:
#         print(f"Excelファイルの読み込み中にエラーが発生しました: {e}")
#         print("ファイル形式が間違っている可能性があります（.xlsではなく.xlsxである必要があります）。")
#         return {}
    
#     return extracted_data_by_sheet

# # --- メインの実行ブロック ---
# if __name__ == "__main__":
#     # --- 1. 変換したい.xlsファイルのパスを設定 ---
#     input_xls_file_path = '/Users/nagaokashuuhei/Desktop/sys_practice/PISsys/history_200 (1).xls'

#     # --- 2. 変換後の.xlsxファイルを保存するフォルダを設定 (任意) ---
#     output_directory_for_conversion = '/Users/nagaokashuuhei/Desktop/sys_practice/PISsys。' 

#     # --- 3. 変換後の.xlsxファイルから抽出したいセルのペアを辞書形式で設定 ---
#     # キーと値がそれぞれ抽出したいセルアドレスになります。
#     # 例: 'A8'の値を抽出し、それに対応する'V8'の値も抽出する
#     target_cell_pairs_to_extract = {
#         'A8': 'V8',
#         'A10': 'V10',
#         'A12': 'V12' # 別のペアの例
#     } 

#     # --- 変換処理の実行 ---
#     converted_xlsx_path = convert_xls_to_xlsx(input_xls_file_path, output_directory_for_conversion)

#     if converted_xlsx_path:
#         print(f"\n変換されたExcelファイル: {converted_xlsx_path}")
#         print("\n--- 変換されたファイルからのセル抽出を開始します ---")
        
#         # --- 抽出処理の実行 ---
#         all_sheets_cells_data = extract_specified_cell_pairs_from_all_sheets(
#             converted_xlsx_path, 
#             target_cell_pairs_to_extract
#         )

#         if all_sheets_cells_data:
#             print("\n--- 全シートからの複数セルペア抽出結果 ---")
#             for sheet, cells_data in all_sheets_cells_data.items():
#                 print(f"シート '{sheet}':")
#                 for cell_addr, value in cells_data.items():
#                     print(f"  セル '{cell_addr}' の値 -> {value}")
#                 print("-" * 20) 
#         else:
#             print("\nセルの抽出に失敗しました。変換されたファイルの内容を確認してください。")
#     else:
#         print("\nファイルの変換に失敗したため、セル抽出処理をスキップします。")



def convert_xls_to_xlsx(input_file, output_file):
    """
    .xlsファイルを.xlsx形式に変換する関数
    """
    try:
        # .xlsファイルを読み込む
        book = p.get_book(file_name=input_file)
        
        # .xlsx形式で保存する
        book.save_as(output_file)
        
        print(f"'{input_file}' を '{output_file}' に変換しました。")
    except Exception as e:
        print(f"変換中にエラーが発生しました: {e}")

# 実行例
input_xls_file = '/Users/nagaokashuuhei/Desktop/sys_practice/PISsys/history_200 (3).xls'
output_xlsx_file = '/Users/nagaokashuuhei/Desktop/sys_practice/PISsys/history_200 (3).xlsx'



# 実際のファイルパスに置き換えてください
# 存在しないファイルだとエラーになるので注意


# def convert_xls_to_xlsx_in_directory(directory_path):
#     """
#     指定されたディレクトリ内の .xls ファイルを .xlsx に変換する関数
    
#     Args:
#         directory_path (str): 検索対象のディレクトリパス
#     """
#     if not os.path.isdir(directory_path):
#         print(f"エラー: ディレクトリ '{directory_path}' が見つかりません。")
#         return

#     # 指定されたディレクトリ内のファイルをリストアップ
#     for filename in os.listdir(directory_path):
#         # ファイルのフルパスを作成
#         filepath = os.path.join(directory_path, filename)
        
#         # ファイルであり、かつ .xls 拡張子を持つか確認
#         if os.path.isfile(filepath) and filename.endswith('.xls'):
#             try:
#                 # 新しいファイル名 (拡張子を .xlsx に変更)
#                 output_filename = filename.replace('.xls', '.xlsx')
#                 output_filepath = os.path.join(directory_path, output_filename)
                
#                 # pyexcel を使って変換を実行
#                 book = p.get_book(file_name=filepath)
#                 book.save_as(output_filepath)
                
#                 print(f"変換成功: '{filename}' -> '{output_filename}'")
                
#             except Exception as e:
#                 print(f"変換失敗: '{filename}' - エラー: {e}")

# # === 使用例 ===
# # 変換したいファイルがあるディレクトリのパスを指定してください
# # 例: 'C:/Users/ユーザー名/Downloads' や '/Users/ユーザー名/Desktop'
# directory_to_convert = '/Users/nagaokashuuhei/Desktop/sys_practice/PISsys' # ここを変更してください

# convert_xls_to_xlsx_in_directory(directory_to_convert)


def test(filepath,name_cell,date_column,record_column):
    try:
        wb = openpyxl.load_workbook(filepath, data_only = True)
        sheet_1 = wb['施術履歴']
        name = sheet_1[name_cell].value

        result_dict = {}


        for sheet_name in wb.sheetnames:
            date_recoord_dict = {}
            sheet = wb[sheet_name]

            date_col_index = openpyxl.utils.column_index_from_string(date_column)
            record_col_index = openpyxl.utils.column_index_from_string(record_column)
            
            for row in range(8,sheet.max_row + 1,1):
                date_cell = sheet.cell(row = row,column = date_col_index)
                record_cell = sheet.cell(row = row,column = record_col_index)

                if date_cell.value is not None and record_cell.value is not None and str(record_cell.value).startswith('#') != True:
                    date_key = date_cell.value
                    date_recoord_dict[date_key] = record_cell.value

            result_dict[sheet_name] = {
                '氏名' : name ,
                'データ' : date_recoord_dict
            }
        print(result_dict)


    except Exception as e:
        raise
    else:
        pass
    finally:
        pass


paitiant_file = '/Users/nagaokashuuhei/Desktop/sys_practice/PISsys/paitiants_info/history_200 (3).xlsx'
test(paitiant_file,'M5','A','V')