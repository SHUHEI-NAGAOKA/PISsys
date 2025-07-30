import time
import chromedriver_binary
import os
import requests
import shutil
import pyexcel as p
import openpyxl
import google.generativeai as genai
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException # TimeoutException をインポート


#ログインを実行する関数
def login(user_id,kyoten_id,password,taion):
    #ブラウザを開いて、KEiROWのサイトにアクセス
    global driver 
    driver = webdriver.Chrome()
    driver.get("https://mobile.keirow.com/qr_login.php#")


    #最初のページの読み込みを待って、users_id要素が確認できたら次のステップに進む
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, 'users_id'))
        )
        print("ユーザーIDフィールドの読み込みに成功しました。")
    except TimeoutException:
        print("ユーザーIDフィールドの読み込みに失敗しました。サイトURLを確認してください。")
        driver.quit()
        exit()


    #ログインに必要なユーザーID、拠点ID、パスワード、体温、ログインボタンの要素を取得
    users_id_field = driver.find_element(By.NAME, 'users_id') #ユーザーIDの要素を取得
    kyoten_id_field = driver.find_element(By.NAME, 'kyoten_id') #拠点IDの要素を取得
    password_field = driver.find_element(By.NAME, 'password') #パスワードの要素を取得
    taion_field = driver.find_element(By.NAME, 'taion') #体温入力の要素を取得
    select_taion = Select(taion_field) #体温入力のフィールドをセレクト可能なオブジェクトにする
    element = driver.find_element(By.LINK_TEXT, "日報ログイン") #ログインボタンの要素を取得


    #それぞれの要素に対して引数のログイン情報を入力
    users_id_field.send_keys(user_id)
    kyoten_id_field.send_keys(kyoten_id)
    password_field.send_keys(password)
    select_taion.select_by_value(taion)


    # ログインボタンをクリック
    element.click()


    #現在のウィンドウの識別子を変数に格納
    original_window = driver.current_window_handle
    try:
        WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
        all_window_handles = driver.window_handles
        new_window_handle = None

        for handle in all_window_handles:
            if handle != original_window:
                new_window_handle = handle
                break
        #ログインが成功して新しいタブが開いたらそちらに切り替え
        if new_window_handle:
            driver.switch_to.window(new_window_handle)

        else:
            raise Exception("新しいウィンドウハンドルが見つかりませんでした。")

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'メインメニュー')]"))
        )

        #施術履歴のページに遷移
        rireki = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(@href, '/search.do') and contains(., '施術履歴を印刷する')]"))
        )
        rireki.click()

    except Exception as e:
        print(f"要素の検索またはクリック中にエラーが発生しました: {e}")
        screenshot_path = "error_screenshot_after_login_attempt.png"
        driver.save_screenshot(screenshot_path)
        print(f"エラー時のスクリーンショットを '{screenshot_path}' に保存しました。")
        print("--- 最終的なページのソースコード（新しいタブに切り替え後の場合） ---")
        print(driver.page_source)
        print("--------------------------")
        driver.quit()
        exit()


def get_paitiants_info(paitiants_list):
    if type(paitiants_list) is list:
        for paitiant in paitiants_list:
            try:
                paitiant_info = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, f"//a[contains(@href, '/printOnly.do') and contains(., '{paitiant}')]")))
            except Exception as e:
                print("患者情報が見つかりませんでした")

            paitiant_info.click()
            time.sleep(2)

            #後々タプルで日付を渡して簡略化した処理をするようにする。
            year_start = driver.find_element(By.NAME, 'optYearStart')
            month_start = driver.find_element(By.NAME, 'optMonthStart')
            day_start = driver.find_element(By.NAME, 'optDayStart')

            year_end = driver.find_element(By.NAME, 'optYearEnd')
            month_end = driver.find_element(By.NAME, 'optMonthEnd')
            day_end = driver.find_element(By.NAME, 'optDayEnd')

            select_year_start = Select(year_start)
            select_year_start.select_by_value(input_year_start)
            select_month_start = Select(month_start)
            select_month_start.select_by_value(input_month_start)
            select_day_start = Select(day_start)
            select_day_start.select_by_value(input_day_start)

            select_year_end = Select(year_end)
            select_year_end.select_by_value(input_year_end)
            select_month_end = Select(month_end)
            select_month_end.select_by_value(input_month_end)
            select_day_end = Select(day_end)
            select_day_end.select_by_value(input_day_end)

            output_btn = driver.find_element(By.NAME, 'excel')
            output_btn.click()

            time.sleep(3)
            back = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(@href, '/search.do') and contains(., '検索')]")))
            back.click()
            time.sleep(3)


    else:
        print("リストが引数に渡されていません")


def move_file(paitiants_list,file_list):
    files = []
    for entry in os.listdir(download_folder):
        full_path = os.path.join(download_folder, entry)
        if os.path.isfile(full_path):
            files.append(full_path)

    for paitiant in reversed(paitiants_list):
        latest_file = max(files, key=os.path.getmtime)
        file_name = os.path.basename(latest_file)
        source_path = latest_file
        destination_path = os.path.join(destination_folder, file_name)

        try:
            shutil.move(source_path, destination_path)
        except Exception as e:
            print('ファイルの移動中にエラーが発生しました')

        file_list.append(destination_path)

        files.remove(latest_file)


def convert_xls_to_xlsx_in_directory(directory_path):
    """
    指定されたディレクトリ内の .xls ファイルを .xlsx に変換する関数
    
    Args:
        directory_path (str): 検索対象のディレクトリパス
    """
    if not os.path.isdir(directory_path):
        print(f"エラー: ディレクトリ '{directory_path}' が見つかりません。")
        return

    # 指定されたディレクトリ内のファイルをリストアップ
    for filename in os.listdir(directory_path):
        # ファイルのフルパスを作成
        filepath = os.path.join(directory_path, filename)
        
        # ファイルであり、かつ .xls 拡張子を持つか確認
        if os.path.isfile(filepath) and filename.endswith('.xls'):
            try:
                # 新しいファイル名 (拡張子を .xlsx に変更)
                output_filename = filename.replace('.xls', '.xlsx')
                output_filepath = os.path.join(directory_path, output_filename)
                
                # pyexcel を使って変換を実行
                book = p.get_book(file_name=filepath)
                book.save_as(output_filepath)
                
                print(f"変換成功: '{filename}' -> '{output_filename}'")
                os.remove(filepath)
                
            except Exception as e:
                print(f"変換失敗: '{filename}' - エラー: {e}")



def scan_paitiant_info(file_path,name_cell,date_column,record_column,append_file):
    try:
        wb = openpyxl.load_workbook(file_path, data_only = True)
        sheet_1 = wb['施術履歴']
        name = sheet_1[name_cell].value

        result_dict = {}

        for sheet_name in wb.sheetnames:
            date_recoord_dict = {}
            sheet = wb[sheet_name]

            date_col_index = openpyxl.utils.column_index_from_string(date_column)
            record_col_index = openpyxl.utils.column_index_from_string(record_column)

            for row in range(8,sheet.max_row + 1,1):
                date_cell = sheet.cell(row = row,column = date_col_index)
                record_cell = sheet.cell(row = row,column = record_col_index)

                if date_cell.value is not None and record_cell.value is not None and str(record_cell.value).startswith('#') != True:
                    date_key = date_cell.value
                    date_recoord_dict[date_key] = record_cell.value

            result_dict[sheet_name] = {
                '氏名' : name ,
                'データ' : date_recoord_dict
            }

        append_file.append(result_dict)


    except Exception as e:
        print('ファイルの読み取り時にエラーが発生しました')
        print(e)


#実行ブロック--------------------------------------------

#ユーザーの情報をこちらに入力
user_id = "10"
kyoten_id = "t22wj"
password = "10"
taion = "36.4" #あとで体温をランダムに変更する処理もする

paitiants_list = ['岡本','手塚']

input_year_start = "2025"
input_month_start = "6"
input_day_start = "1"

input_year_end = "2025"
input_month_end = "6"
input_day_end = "31"


destination_folder = '/Users/nagaokashuuhei/Desktop/sys_practice/PISsys/paitiants_info'
#ダウンロードフォルダのパスを変数に格納
download_folder = os.path.join(os.path.expanduser('~'), 'Downloads')
translate_files = []
name_cell = 'M5'
date_column = 'A'
record_column = 'V'
send_api_text = []

#geminiのAPIキー
API_KEY = "AIzaSyBoGWeZNwI7emgNasTDu5CZXeTezLNxliA"

login(user_id,kyoten_id,password,taion)
get_paitiants_info(paitiants_list)
driver.quit()
move_file(paitiants_list,translate_files)
convert_xls_to_xlsx_in_directory(destination_folder)
for xlsx_file in os.listdir(destination_folder):
    file_full_path = os.path.join(destination_folder,xlsx_file)
    scan_paitiant_info(file_full_path,name_cell,date_column,record_column,send_api_text)
print(send_api_text)